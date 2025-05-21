import logging
import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.client.default import DefaultBotProperties
from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.base import StorageKey
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
from decouple import config
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# FSM для состояний опроса и пользователя (т е конечная машина состояний)
class AdminStates(StatesGroup):
    WAITING_FOR_TITLE = State()
    WAITING_FOR_QUESTIONS = State()

class UserStates(StatesGroup):
    WAITING_FOR_START = State()
    ANSWERING_QUESTIONS = State()

# Logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Config
ADMINS = config("ADMS").split(",")
TOKEN = config("TKN")
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# Runtime state
user_infos = []  # список словарей: [{username: ..., fio: ...}]
survey_title = ""  # Название опроса
prepared_questions = []  # Список подготовленных вопросов: [(тип, вопрос, опции)]
user_progress = {}  # Прогресс пользователя: {username: current_question_index}
user_results = {}  # Результаты: {username: [(вопрос, ответ, время)]}
poll_id_to_data = {}  # Данные опроса: {poll_id: (username, question_index, question, options)}
admin_chat_id = None  # ID чата админа, создавшего опрос
users_completed = set()  # Пользователи, завершившие опрос
users_total = 0  # Общее количество пользователей

def is_admin(message: types.Message):
    logger.debug(f"[RIGHTS]Current user: {message.from_user.username}")
    return message.from_user.username in ADMINS

@dp.startup()
async def setup_commands(bot: Bot):
    commands = [
        types.BotCommand(command="start", description="Начать создание опроса"),
        types.BotCommand(command="poll", description="Добавить вопрос с вариантами"),
        types.BotCommand(command="text", description="Добавить текстовый вопрос"),
        types.BotCommand(command="finish", description="Завершить создание и начать опрос"),
        types.BotCommand(command="status", description="Проверить статус опроса"),
    ]
    await bot.set_my_commands(commands)

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    global admin_chat_id
    
    if is_admin(message):
        admin_chat_id = message.chat.id
        await state.set_state(AdminStates.WAITING_FOR_TITLE)
        await message.reply(
            "👋 Привет, администратор!\n\n"
            "🔹 Этот бот позволяет проводить опросы среди пользователей Telegram.\n"
            "📝 Сначала введите название опроса:"
        )
    else:
        # Обычный пользователь
        await message.reply(
            "👋 Привет!\n\n"
            "🔹 Я бот-опросник. Когда администратор начнет опрос, "
            "я пришлю вам уведомление."
        )

@dp.message(AdminStates.WAITING_FOR_TITLE)
async def process_title(message: types.Message, state: FSMContext):
    global survey_title, prepared_questions
    
    if not is_admin(message):
        return
    
    # Сбрасываем данные предыдущего опроса
    survey_title = message.text.strip()
    prepared_questions = []
    user_results.clear()
    user_progress.clear()
    users_completed.clear()
    
    await state.set_state(AdminStates.WAITING_FOR_QUESTIONS)
    await message.reply(
        f"✅ Название опроса: <b>{survey_title}</b>\n\n"
        "📤 Теперь пришлите Excel-файл (.xlsx) с Telegram ID в первом столбце и ФИО во втором (опционально).\n"
        "🗳 После загрузки вы сможете добавлять вопросы:\n"
        "  — /poll — вопрос с вариантами\n"
        "  — /text — открытый текстовый вопрос\n"
        "📊 Для завершения подготовки и начала опроса — /finish\n\n"
        "👉 Telegram ID можно получить через бота @username_to_id_bot"
    )

@dp.message(Command("status"))
async def check_status(message: types.Message):
    if not is_admin(message):
        await message.reply("❌ У вас нет прав для этой команды.")
        return
    
    if not user_infos:
        await message.reply("⚠️ Опрос еще не начат или не загружен список пользователей.")
        return
    
    completed = len(users_completed)
    total = len(user_infos)
    remaining = total - completed
    
    completed_list = "\n".join(f"@{username}" for username in users_completed) if users_completed else "пока никто"
    
    await message.reply(
        f"📊 Статус опроса <b>{survey_title}</b>:\n\n"
        f"✅ Завершили: {completed} из {total} ({completed/total*100:.1f}%)\n"
        f"⏳ Осталось: {remaining}\n\n"
        f"👤 Завершившие пользователи:\n{completed_list}"
    )

@dp.message(Command("finish"), AdminStates.WAITING_FOR_QUESTIONS)
async def finish_preparation(message: types.Message, state: FSMContext):
    if not is_admin(message):
        await message.reply("❌ У вас нет прав для этой команды.")
        return
    
    if not user_infos:
        await message.reply("❌ Не загружен список пользователей.")
        return
    
    if not prepared_questions:
        await message.reply("❌ Не добавлено ни одного вопроса.")
        return
    
    global users_total
    users_total = len(user_infos)
    
    # Начинаем рассылку приветствий
    success = 0
    fail = 0
    failed_users = []
    
    for info in user_infos:
        username = info["username"][:-2]
        try:
            chat = await bot.get_chat(username)

            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="OK, начать опрос", callback_data="start_survey")]
            ])

            await bot.send_message(
                chat_id=chat.id,
                text=f"Здравствуйте, я бот-опросник. Пожалуйста, ответьте на несколько вопросов для университета по теме: <b>{survey_title}</b>.",
                reply_markup=keyboard
            )

            # ✅ Получаем FSM-контекст и ставим состояние
            user_fsm = FSMContext(storage=dp.storage, key=StorageKey(bot_id=bot.id, chat_id=chat.id, user_id=chat.id))
            await user_fsm.set_state(UserStates.WAITING_FOR_START)

            user_progress[username] = 0
            success += 1

        except Exception as e:
            logger.warning(f"Не отправлено приветствие @{username}: {e}")
            failed_users.append(username)
            fail += 1
    
    fail_list = "\n".join(f"@{u}" for u in failed_users) if failed_users else "нет"
    await message.reply(
        f"✅ Опрос <b>{survey_title}</b> запущен!\n"
        f"✅ Приветствие отправлено: {success}\n"
        f"⚠️ Ошибок: {fail}\n\n"
        f"❌ Не удалось отправить приветствие:\n{fail_list}\n\n"
        f"Используйте /status для проверки прогресса опроса."
    )
    
    await state.clear()  # Сбрасываем состояние админа

@dp.callback_query(F.data == "start_survey")
async def on_start_survey(callback: types.CallbackQuery, state: FSMContext):
    user_id = str(int(callback.from_user.id))  # ключ в user_progress
    username = callback.from_user.username

    logger.debug(f"[START_SURVEY]User ID: {user_id}, username: {username}")
    print("ALL ANSWERS:", user_progress)
    
    if user_id not in user_progress:
        await callback.message.edit_text("К сожалению, этот опрос уже не активен.")
        return

    await callback.message.edit_text(callback.message.text)

    await state.set_state(UserStates.ANSWERING_QUESTIONS)

    await send_next_question(callback.message.chat.id, user_id)  # передаём user_id
    await callback.answer()

async def send_next_question(chat_id, user_id: str):
    question_index = user_progress[user_id]

    # Проверяем, закончились ли вопросы
    if question_index >= len(prepared_questions):
        await bot.send_message(
            chat_id=chat_id,
            text="✅ Спасибо! Вы ответили на все вопросы опроса."
        )
        
        # Отмечаем пользователя как завершившего опрос
        users_completed.add(user_id)
        
        # Проверяем, все ли пользователи завершили опрос
        if len(users_completed) == users_total and admin_chat_id:
            await send_results_to_admin()
        
        return

    question_type, question_text, options = prepared_questions[question_index]

    # Убираем префикс "Вопрос: " если он есть
    if question_text.startswith("Вопрос:"):
        question_text = question_text[8:].strip()

    if question_type == "poll":
        poll = await bot.send_poll(
            chat_id=chat_id,
            question=question_text,
            options=options,
            is_anonymous=False
        )
        poll_id_to_data[poll.poll.id] = (user_id, question_index, question_text, options)

    elif question_type == "text":
        # Ищем fio по user_id, а не по username
        fio = next(
            (user["fio"] for user in user_infos if str(user.get("user_id")) == user_id),
            None
        )
        greeting = f"{fio}, " if fio else ""
        await bot.send_message(
            chat_id=chat_id,
            text=f"✍️ {greeting}{question_text}"
        )

@dp.poll_answer()
async def handle_poll_answer(poll: types.PollAnswer):
    if poll.poll_id not in poll_id_to_data:
        return
    
    username, question_index, question, options = poll_id_to_data[poll.poll_id]
    answer = options[poll.option_ids[0]] if poll.option_ids else "Без ответа"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Сохраняем ответ
    if username not in user_results:
        user_results[username] = []
    
    user_results[username].append((question, answer, timestamp))
    logger.info(f"{username} → '{answer}' на '{question}'")
    
    # Увеличиваем индекс вопроса
    user_progress[username] += 1
    
    # Получаем чат пользователя
    try:
        chat = await bot.get_chat(username)
        await send_next_question(chat.id, username)
    except Exception as e:
        logger.error(f"Ошибка при отправке следующего вопроса для @{username}: {e}")

@dp.message(UserStates.ANSWERING_QUESTIONS)
async def handle_text_answer(message: types.Message, state: FSMContext):
    username = message.from_user.username
    
    if username not in user_progress:
        return
    
    question_index = user_progress[username]
    
    # Проверяем, находится ли пользователь в процессе ответа на вопросы
    if question_index < len(prepared_questions):
        question_type, question, _ = prepared_questions[question_index]
        
        if question_type == "text":
            # Сохраняем ответ на текстовый вопрос
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if username not in user_results:
                user_results[username] = []
            
            user_results[username].append((question, message.text.strip(), timestamp))
            logger.info(f"{username} → '{message.text.strip()}' на '{question}'")
            
            # Увеличиваем индекс вопроса
            user_progress[username] += 1
            
            # Отправляем следующий вопрос
            await send_next_question(message.chat.id, username)

@dp.message(F.document, AdminStates.WAITING_FOR_QUESTIONS)
async def handle_excel(message: types.Message):
    if not is_admin(message):
        await message.reply("❌ У вас нет прав для этой команды.")
        return
    
    global user_infos
    doc = message.document
    if not doc.file_name.endswith(".xlsx"):
        await message.reply("❌ Пожалуйста, пришли .xlsx файл.")
        return
    
    file = await bot.download(doc)
    wb = load_workbook(file)
    ws = wb.active
    
    user_infos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = str(row[0]).strip().lstrip("@") if row[0] else None
        fio = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if username:
            user_infos.append({"username": username, "fio": fio})
    
    if not user_infos:
        await message.reply("❌ Не удалось загрузить ни одного пользователя. Проверь, что столбец содержит Telegram ID.")
        return
    
    await message.reply(f"✅ Загружено {len(user_infos)} пользователей.\n\nТеперь добавьте вопросы с помощью команд:\n👉 /poll — вопрос с вариантами\n👉 /text — текстовый вопрос")

@dp.message(Command("poll"), AdminStates.WAITING_FOR_QUESTIONS)
async def add_poll_question(message: types.Message):
    if not is_admin(message):
        await message.reply("❌ У вас нет прав для этой команды.")
        return
    
    await message.reply("✅ Добавляем вопрос с вариантами. Отправьте его в формате:\n\nВопрос: Какой ваш любимый цвет?\nКрасный\nСиний\nЗеленый")

@dp.message(Command("text"), AdminStates.WAITING_FOR_QUESTIONS)
async def add_text_question(message: types.Message):
    if not is_admin(message):
        await message.reply("❌ У вас нет прав для этой команды.")
        return
    
    await message.reply("✅ Добавляем текстовый вопрос. Отправьте его в формате:\n\n<code>Вопрос: Ваш вопрос</code>")

@dp.message(F.text.startswith("Вопрос:"), AdminStates.WAITING_FOR_QUESTIONS)
async def process_question(message: types.Message):
    if not is_admin(message):
        return
    
    lines = message.text.strip().split("\n")
    question = lines[0].strip()
    
    if len(lines) > 1:
        # Это вопрос с вариантами
        options = [line.strip() for line in lines[1:] if line.strip()]
        if not options:
            await message.reply("❌ Необходимо указать варианты ответа.")
            return
        
        prepared_questions.append(("poll", question, options))
        await message.reply(f"✅ Добавлен вопрос с вариантами: <b>{question}</b>\nВарианты: {', '.join(options)}\n\nВсего вопросов: {len(prepared_questions)}")
    
    else:
        # Это текстовый вопрос
        prepared_questions.append(("text", question, []))
        await message.reply(f"✅ Добавлен текстовый вопрос: <b>{question}</b>\n\nВсего вопросов: {len(prepared_questions)}")

async def send_results_to_admin():
    if not admin_chat_id:
        logger.error("Нет ID администратора для отправки результатов")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["ID пользователя", "Никнейм", "Вопрос", "Ответ", "Время"])
    
    logger.debug(f"user_results before saving: {user_results}")
    
    for username, answers in user_results.items():
        for answer in answers:
            question, response, timestamp = answer
            fio = next((user["fio"] for user in user_infos if user["username"] == username), "Unknown")
            ws.append([username, fio, question, response, timestamp])
    
    file_path = f"results_{survey_title.replace(' ', '_')}.xlsx"
    wb.save(file_path)
    
    await bot.send_document(
        chat_id=admin_chat_id,
        document=FSInputFile(file_path, filename=f"results_{survey_title}.xlsx"),
        caption=f"📊 Итоги опроса <b>{survey_title}</b> - все {len(users_completed)} пользователей завершили опрос!"
    )
    os.remove(file_path)

async def main():
    logger.info("Bot is starting...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())